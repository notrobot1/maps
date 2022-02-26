from flask import Flask
from flask import request

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import re


app = Flask(__name__)
arr = []
file_path = "gmaps.xlsx"
import os.path
if os.path.exists(file_path):
    print("file done")
else:
    print("file create")
    filepath = "gmaps.xlsx"
    wb = openpyxl.Workbook()
    wb.save(filepath)


@app.route("/")
def hello():
    data = request.args.get('data').split(";;")
    arr.append(data)
    print(data)

    from openpyxl import load_workbook

    workbook_name = 'gmaps.xlsx'
    wb = load_workbook(workbook_name)
    page = wb.active

    # New data to write:
    name = data[0]
    type = data[1]
    number = data[2]


    city = data[5]
    url = data[4]
    # else:
    #     number = data[4]
    #     url = data[3]

    # if re.match('[0-9]*$', data[3].replace(" ","")):
    #     number = data[3]
    # elif re.match('[0-9]*$', data[4].replace(" ","")):
    #     number = data[4]
    # elif re.match('[0-9]*$', data[5].replace(" ", "")):
    #     number = data[5]
    # elif re.match('[0-9]*$', data[6].replace(" ", "")):
    #     number = data[6]






    #new_companies = [[data[0], data[1], data[2], data[3], data[4], data[5]]]
    new_companies = [[name, type, number, city, url]]
    for info in new_companies:
        page.append(info)

    wb.save(filename=workbook_name)

    return "Hello World!"

if __name__ == "__main__":
    app.run()